import io
import re
import base64
import time
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

C_BG_DARK   = "18181C"
C_BG_MID    = "1E1E24"
C_ACCENT    = "C8FF3E"
C_TEXT_DARK = "EDEDEB"
C_TEXT_MID  = "A8A8B0"
C_TEAL      = "3EFFB4"
C_WHITE     = "FFFFFF"
C_HEADER_BG = "26262E"
C_SECTION   = "2C2C36"

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, size=11, color=C_TEXT_DARK, name="Arial"):
    return Font(bold=bold, size=size, color=color, name=name)

def _border(color="2A2A34"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def parse_sections(md: str) -> list:
    clean = re.sub(r'```[\s\S]*?```', '', md).strip()
    clean = re.sub(r'^#{1,3}\s*Final Report\s*$', '', clean, flags=re.MULTILINE).strip()

    sections, current = [], None
    for raw in clean.split('\n'):
        l = raw.strip()
        if not l or l == '---':
            continue
        if re.match(r'^#{1,3}\s', l):
            if current and current['bullets']:
                sections.append(current)
            current = {'title': re.sub(r'^#+\s*', '', l).strip(), 'bullets': []}
        elif re.match(r'^[-*•]\s', l):
            if not current:
                current = {'title': 'Key Findings', 'bullets': []}
            current['bullets'].append(re.sub(r'^[-*•]\s', '', l).strip())
        elif re.match(r'^\d+[.)]\s', l):
            if not current:
                current = {'title': 'Key Findings', 'bullets': []}
            current['bullets'].append(re.sub(r'^\d+[.)]\s', '', l).strip())
        elif l and len(l) > 15:
            if not current:
                current = {'title': 'Summary', 'bullets': []}
            current['bullets'].append(l)

    if current and current['bullets']:
        sections.append(current)

    if not sections:
        sentences = [s.strip() for s in re.split(r'[.!?]', clean) if len(s.strip()) > 10]
        sections = [{'title': 'Key Findings', 'bullets': sentences[:8]}]

    return sections

def _strip_bold(text: str) -> str:
    """Remove **bold** markers, keep plain text."""
    return re.sub(r'\*\*(.+?)\*\*', r'\1', text)

def build_summary_sheet(ws, dataset_name: str, section_count: int, plot_count: int):
    ws.sheet_view.showGridLines = False

    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 52
    ws.row_dimensions[1].height = 8

    ws.merge_cells('B2:C3')
    c = ws['B2']
    c.value          = "Agentic Analyser"
    c.font           = Font(bold=True, size=22, color=C_ACCENT, name="Arial")
    c.fill           = _fill(C_BG_DARK)
    c.alignment      = _center()

    ws.merge_cells('B4:C4')
    c = ws['B4']
    c.value     = "Autonomous Data Intelligence Report"
    c.font      = _font(size=11, color=C_TEXT_MID)
    c.fill      = _fill(C_BG_DARK)
    c.alignment = _center()

    ws.row_dimensions[2].height = 36
    ws.row_dimensions[3].height = 0
    ws.row_dimensions[4].height = 18
    ws.row_dimensions[5].height = 10

    info = [
        ("Dataset",    dataset_name),
        ("Generated",  time.strftime("%d %B %Y, %H:%M")),
        ("Sections",   str(section_count)),
        ("Charts",     str(plot_count)),
    ]
    for i, (label, value) in enumerate(info, start=6):
        row = i + (i - 6)
        ws.row_dimensions[row].height = 22
        lc = ws.cell(row=row, column=2, value=label)
        lc.font      = _font(bold=True, size=10, color=C_ACCENT)
        lc.fill      = _fill(C_BG_MID)
        lc.alignment = _left()
        lc.border    = _border()

        vc = ws.cell(row=row, column=3, value=value)
        vc.font      = _font(size=10, color=C_TEXT_DARK)
        vc.fill      = _fill(C_BG_MID)
        vc.alignment = _left()
        vc.border    = _border()

    for row in ws.iter_rows():
        for cell in row:
            if cell.fill.fgColor.rgb == "00000000":
                cell.fill = _fill(C_BG_DARK)

def build_findings_sheet(ws, sections: list):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 26
    ws.column_dimensions['C'].width = 80
    ws.row_dimensions[1].height = 8

    ws.merge_cells('B2:C2')
    c = ws['B2']
    c.value     = "Analysis Findings"
    c.font      = Font(bold=True, size=16, color=C_ACCENT, name="Arial")
    c.fill      = _fill(C_BG_DARK)
    c.alignment = _center()
    ws.row_dimensions[2].height = 30

    ws.row_dimensions[3].height = 6
    ws.row_dimensions[4].height = 20
    for col, label in [(2, "Section"), (3, "Finding")]:
        c = ws.cell(row=4, column=col, value=label)
        c.font      = Font(bold=True, size=9, color=C_BG_DARK, name="Arial")
        c.fill      = _fill(C_ACCENT)
        c.alignment = _center()
        c.border    = _border(C_ACCENT)

    row = 5
    for section in sections:

        ws.row_dimensions[row].height = 22
        ws.merge_cells(f'B{row}:C{row}')
        c = ws.cell(row=row, column=2)
        c.value     = section['title'].upper()
        c.font      = Font(bold=True, size=10, color=C_TEAL, name="Arial")
        c.fill      = _fill(C_SECTION)
        c.alignment = _left()
        c.border    = _border("3EFFB4")
        row += 1

        for i, bullet in enumerate(section['bullets']):
            ws.row_dimensions[row].height = 42
            plain  = _strip_bold(bullet)

            match  = re.match(r'\*\*(.+?)\*\*[:\s]*(.*)', bullet)
            label  = match.group(1) if match else f"{i+1}."
            detail = match.group(2).strip() if match else plain

            lc = ws.cell(row=row, column=2, value=label)
            lc.font      = Font(bold=True, size=9, color=C_ACCENT, name="Arial")
            lc.fill      = _fill(C_BG_MID if i % 2 == 0 else C_BG_DARK)
            lc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            lc.border    = _border()

            dc = ws.cell(row=row, column=3, value=detail or plain)
            dc.font      = _font(size=9, color=C_TEXT_DARK)
            dc.fill      = _fill(C_BG_MID if i % 2 == 0 else C_BG_DARK)
            dc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            dc.border    = _border()
            row += 1

        ws.row_dimensions[row].height = 6
        for col in [2, 3]:
            ws.cell(row=row, column=col).fill = _fill(C_BG_DARK)
        row += 1

    for r in ws.iter_rows():
        for cell in r:
            if cell.fill.fgColor.rgb == "00000000":
                cell.fill = _fill(C_BG_DARK)

def build_charts_sheet(ws, plots_b64: list):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 3

    ws.merge_cells('B1:D1')
    c = ws['B1']
    c.value     = "Charts & Visualisations"
    c.font      = Font(bold=True, size=16, color=C_ACCENT, name="Arial")
    c.fill      = _fill(C_BG_DARK)
    c.alignment = _center()
    ws.row_dimensions[1].height = 30

    row = 3
    for idx, b64 in enumerate(plots_b64):
        try:
            img_bytes = base64.b64decode(b64)
            img_buf   = io.BytesIO(img_bytes)
            img       = XLImage(img_buf)

            scale = min(600 / (img.width or 600), 1.0)
            img.width  = int((img.width  or 600) * scale)
            img.height = int((img.height or 400) * scale)

            ws.row_dimensions[row].height   = 15
            label_cell = ws.cell(row=row, column=2, value=f"Chart {idx + 1}")
            label_cell.font      = Font(bold=True, size=9, color=C_TEAL, name="Arial")
            label_cell.fill      = _fill(C_SECTION)
            label_cell.alignment = _left()
            row += 1

            img_row = row

            rows_needed = max(1, int(img.height / 15))
            for r in range(img_row, img_row + rows_needed + 2):
                ws.row_dimensions[r].height = 15

            ws.add_image(img, f"B{img_row}")
            row += rows_needed + 3

        except Exception as e:
            ws.cell(row=row, column=2, value=f"Chart {idx+1} — could not embed: {e}")
            row += 2

    for r in ws.iter_rows():
        for cell in r:
            if cell.fill.fgColor.rgb == "00000000":
                cell.fill = _fill(C_BG_DARK)

def generate_excel(report_md: str, plots_b64: list, dataset_name: str) -> bytes:
    sections = parse_sections(report_md)

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Summary"
    build_summary_sheet(ws1, dataset_name, len(sections), len(plots_b64))

    ws2 = wb.create_sheet("Findings")
    build_findings_sheet(ws2, sections)

    if plots_b64:
        ws3 = wb.create_sheet("Charts")
        build_charts_sheet(ws3, plots_b64)

    wb.active = ws1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
