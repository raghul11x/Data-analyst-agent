import base64
import io
import json
import os
import queue
import tempfile
import threading
import time

import requests as req_lib
import pandas as pd
from flask import Flask, Response, jsonify, request, send_from_directory, stream_with_context
from flask_cors import CORS

from database import (
    init_db, create_session, update_final_report, add_plot,
    get_full_session, list_sessions, delete_session, get_dataset_path,
    create_prep_session, update_prep_log, get_prep_session, get_latest_prep_session,
    evict_old_sessions
)

init_db()
evict_old_sessions()
from agent import (
    SUGGEST_SYSTEM,
    call_llm_fast,
    run_agent_stream,
    run_followup_stream,
)
from config import (
    API_KEY, API_URL, MODEL,
    SESSIONS, _session_lock,
    session_plot_path,
)
from pdf import generate_pdf
from excel_export import generate_excel
from utils import load_dataset, load_multiple_datasets, sse
from preprocessor import profile_dataframe, apply_pipeline

BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
STATIC_DIR = os.path.join(BASE_DIR, "static")

app = Flask(__name__, static_folder=STATIC_DIR)
CORS(app)

def _startup_check():
    if not API_KEY:
        print("\n[startup] ✗  OPENROUTER_API_KEY not set — create a .env file")
        return
    print(f"\n[startup] Testing model: {MODEL} …")
    try:
        resp = req_lib.post(
            url=API_URL,
            headers={
                "Authorization": f"Bearer {API_KEY}",
                "Content-Type": "application/json",
                "HTTP-Referer": "http://localhost:5050",
            },
            json={
                "model": MODEL,
                "messages": [{"role": "user", "content": "Say: OK"}],
                "max_tokens": 5,
            },
            timeout=20,
        )
        if resp.status_code == 200:
            print("[startup] ✓  Model ready")
        elif resp.status_code == 429:
            print("[startup] ⚠  Rate limited — wait a moment then refresh")
        else:
            print(f"[startup] ✗  HTTP {resp.status_code}: {resp.text[:100]}")
    except Exception as e:
        print(f"[startup] ✗  Connection error: {e}")
    print()

@app.route("/")
def index():
    """Serve the frontend using an absolute path — fixes 404 in VS Code terminal."""
    html_path = os.path.join(STATIC_DIR, "index.html")
    with open(html_path, encoding="utf-8") as f:
        return f.read(), 200, {"Content-Type": "text/html; charset=utf-8"}

@app.route("/model")
def get_model():
    return jsonify({"model": MODEL})

@app.route("/static/<path:filename>")
def static_files(filename):
    return send_from_directory(STATIC_DIR, filename)

@app.route("/favicon.ico")
def favicon():
    return send_from_directory(STATIC_DIR, "favicon.png", mimetype="image/png")

@app.route("/preprocess")
def preprocess_page():
    html_path = os.path.join(STATIC_DIR, "preprocess.html")
    with open(html_path, encoding="utf-8") as f:
        return f.read(), 200, {"Content-Type": "text/html; charset=utf-8"}

@app.route("/columns", methods=["POST"])
def get_columns():
    if "file" not in request.files:
        return jsonify({"columns": [], "dtypes": {}}), 400
    try:
        df, _ = load_dataset(request.files["file"])
        return jsonify({
            "columns": list(df.columns),
            "dtypes":  {c: str(t) for c, t in df.dtypes.items()},
        })
    except Exception as e:
        return jsonify({"columns": [], "dtypes": {}, "error": str(e)}), 400

@app.route("/suggest", methods=["POST"])
def suggest():
    data    = request.json or {}
    columns = data.get("columns", [])
    dtypes  = data.get("dtypes", {})
    if not columns:
        return jsonify({"questions": []}), 200

    col_info = ", ".join(f"{c} ({dtypes.get(c, '?')})" for c in columns[:20])
    raw = call_llm_fast(
        [{"role": "user",
          "content": f"Dataset columns: {col_info}\n\nSuggest 5 specific analysis questions."}],
        system=SUGGEST_SYSTEM,
    )
    if not raw:
        return jsonify({"questions": []}), 200

    print(f"[suggest] raw response: {raw[:300]}")

    import re as _re

    try:
        arr_match = _re.search(r'\[\s*"[\s\S]*?"\s*\]', raw, _re.DOTALL)
        if arr_match:
            questions = json.loads(arr_match.group())
            if isinstance(questions, list):
                qs = [str(q).strip() for q in questions if str(q).strip()]
                if qs:
                    print(f"[suggest] JSON path: {qs}")
                    return jsonify({"questions": qs[:5]})
    except Exception:
        pass

    try:
        clean = raw.strip().strip("`").strip()
        if clean.lower().startswith("json"):
            clean = clean[4:].strip()
        questions = json.loads(clean)
        if isinstance(questions, list):
            qs = [str(q).strip() for q in questions if str(q).strip()]
            if qs:
                print(f"[suggest] clean-JSON path: {qs}")
                return jsonify({"questions": qs[:5]})
    except Exception:
        pass

    lines = []
    for l in raw.split("\n"):

        l = l.strip()
        l = _re.sub(r'^[\-\*•]\s*', '', l)
        l = _re.sub(r'^\d+[\.):]\s*', '', l)
        l = l.strip('"\'` \t,')
        if len(l) > 15 and "?" in l:
            lines.append(l)
    print(f"[suggest] fallback path: {lines}")
    return jsonify({"questions": lines[:5]})

@app.route("/analyse", methods=["POST"])
def analyse():
    files   = request.files.getlist("file")
    if not files or not files[0].filename:
        return {"error": "No file uploaded"}, 400

    q_text   = request.form.get("question", "Analyse this dataset fully.")
    max_iter = int(request.form.get("max_iterations", 8))
    sid      = f"s_{int(time.time() * 1000)}"

    try:
        if len(files) == 1:
            df, info = load_dataset(files[0])
            tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".csv")
            df.to_csv(tmp.name, index=False)
            perm_path     = tmp.name
            dataset_paths = {files[0].filename: tmp.name}
            dataset_names = files[0].filename
        else:
            dataset_paths, info = load_multiple_datasets(files)
            dataset_names = ", ".join(dataset_paths.keys())
            perm_path     = list(dataset_paths.values())[0]
    except Exception as e:
        return {"error": str(e)}, 400

    with _session_lock:
        SESSIONS[sid] = {
            "messages":      [],
            "dataset_info":  info,
            "dataset_paths": dataset_paths,
            "dataset_path":  perm_path,
            "dataset_name":  dataset_names,
            "plot_path":     session_plot_path(sid),
            "plots":         [],
            "final_report":  None,
            "created_at":    time.time(),
        }

    q = queue.Queue()
    threading.Thread(
        target=run_agent_stream,
        args=(sid, dataset_paths, info, q_text, max_iter, q),
        daemon=True,
    ).start()

    def generate():
        yield sse("session", {"session_id": sid})
        while True:
            item = q.get()
            if item is None:
                break
            yield item

        _save_session(sid)

    return Response(
        stream_with_context(generate()),
        mimetype="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )

@app.route("/followup", methods=["POST"])
def followup():
    data   = request.json or {}
    sid    = data.get("session_id", "")
    q_text = data.get("question", "").strip()

    if not sid or sid not in SESSIONS:
        return {"error": "Session not found. Run an analysis first."}, 400
    if not q_text:
        return {"error": "No question provided."}, 400

    q = queue.Queue()
    threading.Thread(
        target=run_followup_stream,
        args=(sid, q_text, q),
        daemon=True,
    ).start()

    def generate():
        while True:
            item = q.get()
            if item is None:
                break
            yield item

    return Response(
        stream_with_context(generate()),
        mimetype="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no", "Connection": "keep-alive"},
    )

@app.route("/export_pdf", methods=["POST"])
def export_pdf():
    data = request.json or {}
    sid  = data.get("session_id", "")

    if not sid or sid not in SESSIONS:
        return {"error": "Session not found."}, 400

    session = SESSIONS.get(sid, {})
    report  = session.get("final_report", "")

    if not report:
        db_session = get_full_session(sid)
        if db_session:
            report = db_session.get("final_report", "")
            session = db_session

    if not report:
        return {"error": "No final report yet. Run an analysis first."}, 400

    all_plots = get_full_session(sid)
    plots = all_plots.get("plots", []) if all_plots else session.get("plots", [])

    try:
        pdf  = generate_pdf(
            report,
            plots,
            session.get("dataset_name", session.get("name", "dataset")),
        )
        name = session.get("dataset_name", "dataset").rsplit(".", 1)[0]
        return jsonify({
            "pdf":      base64.b64encode(pdf).decode(),
            "filename": f"analysis_{name}.pdf",
        })
    except Exception as e:
        return {"error": f"PDF generation failed: {e}"}, 500

PREP_SESSIONS = {}

def _read_any_file(file_storage):
    """Read CSV or Excel into a DataFrame as fast as possible.
    Always tries CSV first (sniff content), then falls back to Excel engines.
    """
    fname = file_storage.filename.lower()

    suffix = ".csv" if fname.endswith(".csv") else (".xlsx" if fname.endswith(".xlsx") else ".xls")
    raw = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
    try:
        file_storage.save(raw.name)
        raw.close()

        try:
            df = pd.read_csv(raw.name, low_memory=False)
            if len(df.columns) > 1:
                return df
        except Exception:
            pass

        for engine in ["calamine", "openpyxl", "xlrd", None]:
            try:
                kwargs = {"engine": engine} if engine else {}
                df = pd.read_excel(raw.name, **kwargs)
                return df
            except Exception:
                continue

        raise ValueError("Could not read file with any supported format.")
    finally:
        try:
            os.unlink(raw.name)
        except Exception:
            pass

@app.route("/export_excel", methods=["POST"])
def export_excel():
    data = request.json or {}
    sid  = data.get("session_id", "")
    if not sid or sid not in SESSIONS:
        return {"error": "Session not found."}, 400

    session = SESSIONS.get(sid, {})
    report  = session.get("final_report", "")
    if not report:

        try:
            db_s = get_full_session(sid)
            if db_s:
                report = db_s.get("final_report", "")
                session = db_s
        except Exception:
            pass
    if not report:
        return {"error": "No final report yet."}, 400

    try:
        all_plots = get_full_session(sid)
        plots = all_plots.get("plots", []) if all_plots else session.get("plots", [])
        name  = session.get("dataset_name", "dataset").rsplit(".", 1)[0]
        xlsx  = generate_excel(report, plots, session.get("dataset_name", "dataset"))
        return jsonify({
            "excel":    base64.b64encode(xlsx).decode(),
            "filename": f"analysis_{name}.xlsx",
        })
    except Exception as e:
        return {"error": f"Excel export failed: {e}"}, 500

@app.route("/prep/profile", methods=["POST"])
def prep_profile():
    """Upload a file, convert to CSV internally, return full column profile."""
    if "file" not in request.files:
        return jsonify({"error": "No file"}), 400
    file = request.files["file"]
    try:
        df = _read_any_file(file)
    except Exception as e:
        return jsonify({"error": f"Could not read file: {e}"}), 400

    sid = f"prep_{int(time.time()*1000)}"
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".csv")

    df.to_csv(tmp.name, index=False)
    PREP_SESSIONS[sid] = {
        "path":       tmp.name,
        "filename":   file.filename,
        "log":        [],
        "created_at": time.time(),
    }
    create_prep_session(sid, file.filename, tmp.name)

    profile = profile_dataframe(df)
    profile["session_id"]   = sid
    profile["preview"]      = df.head(5).fillna("").astype(str).to_dict(orient="records")
    profile["columns_list"] = list(df.columns)
    return jsonify(profile)

@app.route("/prep/apply", methods=["POST"])
def prep_apply():
    """Apply a preprocessing pipeline and return updated profile + preview."""
    data     = request.json or {}
    sid      = data.get("session_id", "")
    pipeline = data.get("pipeline", {})

    if sid not in PREP_SESSIONS:
        return jsonify({"error": "Session not found"}), 400

    path = PREP_SESSIONS[sid]["path"]
    try:
        df = pd.read_csv(path)
        cleaned_df, log, code = apply_pipeline(df, pipeline)

        cleaned_df.to_csv(path, index=False)
        PREP_SESSIONS[sid]["log"].extend(log)
        PREP_SESSIONS[sid]["code"] = code
        update_prep_log(sid, PREP_SESSIONS[sid]["log"], code)

        profile = profile_dataframe(cleaned_df)
        profile["session_id"] = sid
        profile["log"]        = log
        profile["preview"]    = cleaned_df.head(5).fillna("").astype(str).to_dict(orient="records")
        profile["columns_list"] = list(cleaned_df.columns)
        return jsonify(profile)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/prep/download_csv", methods=["POST"])
def prep_download_csv():
    """Return the cleaned CSV as base64."""
    data = request.json or {}
    sid  = data.get("session_id", "")
    if sid not in PREP_SESSIONS:
        return jsonify({"error": "Session not found"}), 400
    path = PREP_SESSIONS[sid]["path"]
    with open(path, "rb") as f:
        b64 = base64.b64encode(f.read()).decode()
    orig = PREP_SESSIONS[sid].get("filename", "dataset.csv")
    name = orig.rsplit(".", 1)[0] + "_cleaned.csv"
    return jsonify({"csv": b64, "filename": name})

@app.route("/prep/export_xlsx", methods=["POST"])
def prep_export_xlsx():
    """Convert a CSV file to formatted xlsx and return as base64."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    if "file" not in request.files:
        return jsonify({"error": "No file"}), 400
    file     = request.files["file"]
    basename = request.form.get("filename", "cleaned_dataset")
    try:
        df = pd.read_csv(file, low_memory=False)
    except Exception as e:
        return jsonify({"error": str(e)}), 400

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    accent = "C8FF3E"
    for ci, col in enumerate(df.columns, 1):
        c = ws.cell(row=1, column=ci, value=col)
        c.font      = Font(bold=True, color="111111", name="Arial", size=10)
        c.fill      = PatternFill("solid", fgColor=accent)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = max(12, min(len(str(col)) + 4, 30))

    for ri, row in enumerate(df.itertuples(index=False), 2):
        bg = "1E1E24" if ri % 2 == 0 else "26262E"
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val if not pd.isna(val) else "")
            c.font      = Font(color="EDEDEB", name="Arial", size=9)
            c.fill      = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal="left", vertical="center")

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return jsonify({
        "xlsx":     base64.b64encode(buf.read()).decode(),
        "filename": basename + ".xlsx"
    })

@app.route("/prep/download_code", methods=["POST"])
def prep_download_code():
    """Return the preprocessing Python script as base64."""
    data = request.json or {}
    sid  = data.get("session_id", "")
    if sid not in PREP_SESSIONS:
        return jsonify({"error": "Session not found"}), 400
    code = PREP_SESSIONS[sid].get("code", "# No pipeline applied yet")
    b64  = base64.b64encode(code.encode()).decode()
    return jsonify({"code": b64, "filename": "preprocessing.py"})

@app.route("/prep/pending", methods=["GET"])
def prep_pending():
    """Return the most recent prep session that has been processed."""
    if not PREP_SESSIONS:
        return jsonify({"pending": False})
    candidates = [(sid, s) for sid, s in PREP_SESSIONS.items() if s.get("log")]
    if not candidates:
        return jsonify({"pending": False})
    sid, session = max(candidates, key=lambda x: x[1].get("created_at", 0))
    return jsonify({
        "pending":    True,
        "session_id": sid,
        "filename":   session.get("filename", "cleaned_dataset.csv"),
    })

@app.route("/prep/get_csv/<session_id>", methods=["GET"])
def prep_get_csv(session_id):
    """Stream the cleaned CSV file directly."""
    if session_id not in PREP_SESSIONS:
        return {"error": "Session not found"}, 404
    path = PREP_SESSIONS[session_id]["path"]
    filename = PREP_SESSIONS[session_id].get("filename", "cleaned.csv")
    name = filename.rsplit(".", 1)[0] + "_cleaned.csv"
    with open(path, "rb") as f:
        data = f.read()
    from flask import make_response
    resp = make_response(data)
    resp.headers["Content-Type"] = "text/csv"
    resp.headers["Content-Disposition"] = f"attachment; filename={name}"
    return resp

@app.route("/prep/full_log", methods=["POST"])
def prep_full_log():
    data = request.json or {}
    sid  = data.get("session_id", "")
    if sid not in PREP_SESSIONS:
        return jsonify({"error": "Session not found"}), 400
    return jsonify({"log": PREP_SESSIONS[sid].get("log", [])})

import json as _json

SESSIONS_DIR  = os.path.join(BASE_DIR, ".sessions")
DATASETS_DIR  = os.path.join(BASE_DIR, ".datasets")
os.makedirs(SESSIONS_DIR, exist_ok=True)
os.makedirs(DATASETS_DIR, exist_ok=True)

def _save_session(sid: str):
    """Save completed session to DB for history."""
    session = SESSIONS.get(sid)
    if not session or not session.get("final_report"):
        return
    try:
        from database import add_plot as _ap, update_final_report as _ur
        create_session(sid, session.get("dataset_name",""), session.get("dataset_path",""), "")
        for plot in session.get("plots", []):
            _ap(sid, plot)
        _ur(sid, session["final_report"])
    except Exception as e:
        print(f"[db] save error: {e}")

@app.route("/analyse/get_csv/<session_id>", methods=["GET"])
def analyse_get_csv(session_id):
    """Serve the original dataset — checks memory first, then disk."""
    path     = None
    filename = "dataset.csv"

    if session_id in SESSIONS:
        session  = SESSIONS[session_id]
        path     = session.get("dataset_path", "")
        filename = session.get("dataset_name", "dataset.csv")

    if not path or not os.path.exists(path):
        db_path = get_dataset_path(session_id)
        if db_path and os.path.exists(db_path):
            path = db_path
            db_session = get_full_session(session_id)
            if db_session:
                filename = db_session.get("dataset_name", "dataset.csv")

    if not path or not os.path.exists(path):
        perm_check = os.path.join(DATASETS_DIR, f"{session_id}.csv")
        if os.path.exists(perm_check):
            path = perm_check

    if not path or not os.path.exists(path):
        return jsonify({"error": "File no longer available. Please re-upload."}), 404

    with open(path, "rb") as f:
        data = f.read()
    from flask import make_response
    resp = make_response(data)
    resp.headers["Content-Type"] = "text/csv"
    resp.headers["Content-Disposition"] = f"attachment; filename={filename}"
    return resp

@app.route("/sessions/list", methods=["GET"])
def sessions_list():
    """Return the 10 most recent saved sessions from DB."""
    try:
        rows = list_sessions(10)
        sessions = [{
            "session_id":   r["id"],
            "dataset_name": r["dataset_name"] or "Unknown",
            "created_at":   r["created_at"],
            "has_plots":    r["plot_count"] > 0,
        } for r in rows if r.get("has_report")]
        return jsonify({"sessions": sessions})
    except Exception as e:
        return jsonify({"sessions": [], "error": str(e)})

@app.route("/sessions/load/<session_id>", methods=["GET"])
def sessions_load(session_id):
    """Load a saved session from DB."""
    try:
        data = get_full_session(session_id)
        if not data:
            return jsonify({"error": "Session not found"}), 404
        return jsonify({
            "session_id":   data["id"],
            "dataset_name": data["dataset_name"],
            "final_report": data["final_report"],
            "plots":        data["plots"],
            "created_at":   data["created_at"],
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/sessions/delete/<session_id>", methods=["DELETE"])
def sessions_delete(session_id):
    try:
        delete_session(session_id)
    except Exception:
        pass
    return jsonify({"ok": True})

if __name__ == "__main__":
    os.makedirs(STATIC_DIR, exist_ok=True)
    init_db()
    evict_old_sessions()
    print("""
  ╔══════════════════════════════════════╗
  ║         Agentic Analyser             ║
  ╚══════════════════════════════════════╝
""")
    _startup_check()
    print("  ✦  Open http://localhost:5050\n")

    try:
        from waitress import serve
        print("  Using waitress server (streaming enabled)\n")
        serve(app, host="0.0.0.0", port=5050, threads=8,
              channel_timeout=600, cleanup_interval=30,
              send_bytes=1)
    except ImportError:
        print("  waitress not found — using Flask dev server")
        print("  Run: pip install waitress  for better streaming\n")
        app.run(host="0.0.0.0", port=5050, debug=False, threaded=True)
